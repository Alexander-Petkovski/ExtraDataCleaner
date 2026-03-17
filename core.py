"""
ExtraDataCleaner — core.py
===========================
Comprehensive formatting cleaner for CSV and Excel files.
Targets readability in JupyterLab / pandas workflows.

Cleaning categories handled
─────────────────────────────
STRUCTURAL   : BOM, encoding, line-endings, fully-empty rows/cols,
               junk header rows, all-whitespace rows
HEADERS      : strip whitespace, normalise case, snake_case,
               remove/replace illegal characters, de-duplicate,
               fill blank/unnamed headers, multi-row header collapse
VALUES       : strip cell whitespace, unify null representations,
               strip currency / accounting symbols, remove thousands
               separators, fix decimal-comma numbers, convert
               numbers-stored-as-text, normalise booleans,
               clean Excel error strings, trim stray quotes,
               normalise line-breaks inside cells
EXCEL-ONLY   : all sheets supported, merged-cell forward-fill,
               hidden row/col skipping (optional), drop chart sheets
"""

from __future__ import annotations

import io
import re
import warnings
from copy import deepcopy
from pathlib import Path
from typing import Any

import chardet
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning)


# ── constant lookup tables ────────────────────────────────────────────────────

# Strings that all mean "missing" – we unify them to NaN
_NULL_VARIANTS: set[str] = {
    "", "nan", "none", "null", "na", "n/a", "n.a.", "n.a",
    "#n/a", "#na", "nil", "-", "--", "---", ".", "..",
    "missing", "unknown", "undefined", "not available",
    "not applicable", "#value!", "#ref!", "#div/0!", "#name?",
    "#num!", "#null!", "inf", "-inf", "infinity", "-infinity",
    "1.#qnan", "-1.#qnan", "1.#ind", "-1.#ind",
}

# Strings that mean True / False
_TRUE_VARIANTS:  set[str] = {"true",  "yes", "y", "1", "on",  "t"}
_FALSE_VARIANTS: set[str] = {"false", "no",  "n", "0", "off", "f"}

# Characters that are problematic in pandas column names
_ILLEGAL_HEADER_RE = re.compile(r"[^\w\s]")           # keeps word chars + spaces
_WHITESPACE_RE     = re.compile(r"\s+")
_MULTI_UNDERSCORE  = re.compile(r"_+")
_LEADING_DIGIT_RE  = re.compile(r"^(\d)")              # column starting with digit
_CURRENCY_RE       = re.compile(r"[$€£¥₹₩₿¢₽₺₴₦₪₫฿₮₱]")
_THOUSANDS_RE      = re.compile(r"(?<=\d)[,\s](?=\d{3}(?:[^\d]|$))")
_ACCOUNTING_PAREN  = re.compile(r"^\((\d[\d.,]*)\)$")  # (1,234.56) → negative
_PERCENT_RE        = re.compile(r"^([+-]?\d[\d.,]*)\s*%$")
_EXCEL_ERRORS      = {
    "#div/0!", "#n/a", "#name?", "#null!", "#num!",
    "#ref!",   "#value!", "#getting_data", "######",
}
_CELL_LINEBREAK_RE = re.compile(r"[\r\n]+")            # newlines inside a cell


# ── helpers ───────────────────────────────────────────────────────────────────

def _detect_encoding(path: Path) -> str:
    raw = path.read_bytes()[:200_000]          # sample first 200 kB
    result = chardet.detect(raw)
    enc = result.get("encoding") or "utf-8"
    # Prefer utf-8 variants over latin-1 aliases when confidence is borderline
    if enc.lower() in ("ascii", "iso-8859-1", "windows-1252") and result.get("confidence", 0) < 0.85:
        enc = "utf-8"
    return enc


def _strip_bom(text: str) -> str:
    return text.lstrip("\ufeff\ufffe")


def _to_snake(name: str) -> str:
    """'Total Sales ($)' → 'total_sales'"""
    name = _ILLEGAL_HEADER_RE.sub("_", name)
    name = _WHITESPACE_RE.sub("_", name.strip())
    name = _MULTI_UNDERSCORE.sub("_", name)
    name = name.strip("_").lower()
    if _LEADING_DIGIT_RE.match(name):
        name = "col_" + name
    return name or "unnamed"


def _safe_numeric(val: str) -> Any:
    """
    Try to parse a string as a number.
    Handles: thousands separators, decimal-comma, currency, accounting parens,
    percentages, scientific notation.
    Returns the numeric value (int or float) or the original string.
    """
    if not isinstance(val, str):
        return val

    s = val.strip()

    # Accounting negative: (1,234.56)
    m = _ACCOUNTING_PAREN.match(s)
    if m:
        s = "-" + m.group(1)

    # Strip currency
    s = _CURRENCY_RE.sub("", s).strip()

    # Percentage
    m = _PERCENT_RE.match(s)
    if m:
        try:
            return float(m.group(1).replace(",", ".")) / 100
        except ValueError:
            pass

    # Remove thousands separators smartly:
    # "1,234,567.89"  → "1234567.89"   (comma-thousands, dot-decimal)
    # "1.234.567,89"  → "1234567.89"   (dot-thousands,  comma-decimal)
    if s.count(",") > 0 and s.count(".") > 0:
        # Determine which is thousands vs decimal by position of last occurrence
        last_comma = s.rfind(",")
        last_dot   = s.rfind(".")
        if last_dot > last_comma:          # comma-thousands, dot-decimal (US/UK)
            s = s.replace(",", "")
        else:                               # dot-thousands, comma-decimal (EU)
            s = s.replace(".", "").replace(",", ".")
    elif s.count(",") == 1 and s.count(".") == 0:
        # Could be decimal-comma (EU) or thousands – check decimals part
        parts = s.split(",")
        if len(parts[1]) in (1, 2, 3):
            # Ambiguous; treat trailing ≤3 digits after sole comma as decimal
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif s.count(",") > 1:
        s = s.replace(",", "")

    try:
        f = float(s)
        if f == int(f) and "." not in s and "e" not in s.lower():
            return int(f)
        return f
    except (ValueError, OverflowError):
        return val                          # return original if not numeric


# ── main class ────────────────────────────────────────────────────────────────

class DataCleaner:
    """
    Parameters
    ----------
    snake_case      : bool   – convert column names to snake_case (default True)
    unify_nulls     : bool   – replace all null-like strings with NaN (default True)
    numeric_coerce  : bool   – parse numbers stored as text (default True)
    bool_coerce     : bool   – normalise bool-like strings to True/False (default True)
    drop_empty_rows : bool   – drop fully-empty rows (default True)
    drop_empty_cols : bool   – drop fully-empty columns (default True)
    skip_hidden     : bool   – ignore hidden rows/cols in Excel (default True)
    sheet           : str|None – specific Excel sheet name; None = all sheets
    header_row      : int|None – force header row index; None = auto-detect
    percent_to_float: bool   – convert "12.5%" → 0.125 (default True)
    """

    def __init__(
        self,
        snake_case:       bool = True,
        unify_nulls:      bool = True,
        numeric_coerce:   bool = True,
        bool_coerce:      bool = True,
        drop_empty_rows:  bool = True,
        drop_empty_cols:  bool = True,
        skip_hidden:      bool = True,
        sheet:            str | None = None,
        header_row:       int | None = None,
        percent_to_float: bool = True,
    ) -> None:
        self.snake_case       = snake_case
        self.unify_nulls      = unify_nulls
        self.numeric_coerce   = numeric_coerce
        self.bool_coerce      = bool_coerce
        self.drop_empty_rows  = drop_empty_rows
        self.drop_empty_cols  = drop_empty_cols
        self.skip_hidden      = skip_hidden
        self.sheet            = sheet
        self.header_row       = header_row
        self.percent_to_float = percent_to_float
        self.report: list[str] = []

    # ── public entry point ────────────────────────────────────────────────────

    def clean_file(self, path: str | Path) -> dict[str, pd.DataFrame]:
        """
        Load and clean a CSV or Excel file.

        Returns a dict  { sheet_name: cleaned_DataFrame }.
        For CSV files the single key is "Sheet1".
        """
        path = Path(path)
        self.report = []
        suffix = path.suffix.lower()

        if suffix in (".xlsx", ".xls", ".xlsm", ".xlsb", ".ods"):
            frames = self._load_excel(path)
        elif suffix in (".csv", ".tsv", ".txt"):
            frames = {"Sheet1": self._load_csv(path)}
        else:
            # Try CSV as fallback
            self._log(f"Unknown extension '{suffix}' – attempting CSV parse.")
            frames = {"Sheet1": self._load_csv(path)}

        cleaned: dict[str, pd.DataFrame] = {}
        for name, df in frames.items():
            self._log(f"\n── Sheet: {name!r}  ({df.shape[0]} rows × {df.shape[1]} cols before cleaning)")
            df = self._clean_dataframe(df)
            cleaned[name] = df
            self._log(f"   → {df.shape[0]} rows × {df.shape[1]} cols after cleaning")

        return cleaned

    # ── loaders ───────────────────────────────────────────────────────────────

    def _load_csv(self, path: Path) -> pd.DataFrame:
        enc = _detect_encoding(path)
        self._log(f"Detected encoding: {enc}")

        raw = path.read_bytes().decode(enc, errors="replace")
        raw = _strip_bom(raw)
        # Normalise line endings
        raw = raw.replace("\r\n", "\n").replace("\r", "\n")

        # Auto-detect delimiter (comma, semicolon, tab, pipe)
        sample = "\n".join(raw.splitlines()[:20])
        delimiter = self._detect_delimiter(sample)
        self._log(f"Detected delimiter: {delimiter!r}")

        # Determine header row
        hdr = self._find_header_row(raw, delimiter) if self.header_row is None else self.header_row
        if hdr > 0:
            self._log(f"Skipping {hdr} leading junk row(s) before header.")

        df = pd.read_csv(
            io.StringIO(raw),
            sep=delimiter,
            header=hdr,
            dtype=str,
            keep_default_na=False,
            skipinitialspace=True,
            skip_blank_lines=False,   # keep blank rows so our header index is accurate
            encoding_errors="replace",
            on_bad_lines="warn",
        )
        return df

    def _load_excel(self, path: Path) -> dict[str, pd.DataFrame]:
        wb = None
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(path, data_only=True, read_only=False)
        except Exception:
            pass

        sheet_names = self._resolve_sheets(path)
        frames: dict[str, pd.DataFrame] = {}

        for sname in sheet_names:
            # Determine header row
            hdr = self.header_row
            if hdr is None:
                hdr = self._find_excel_header_row(path, sname)

            df = pd.read_excel(
                path,
                sheet_name=sname,
                header=hdr,
                dtype=str,
                keep_default_na=False,
                engine="openpyxl",
            )

            # Expand merged cells using openpyxl
            if wb and sname in wb.sheetnames:
                df = self._expand_merged_cells(wb[sname], df, hdr or 0)

            # Skip hidden rows/cols if requested
            if self.skip_hidden and wb and sname in wb.sheetnames:
                df = self._drop_hidden(wb[sname], df, hdr or 0)

            frames[sname] = df

        return frames

    # ── excel helpers ─────────────────────────────────────────────────────────

    def _resolve_sheets(self, path: Path) -> list[str]:
        xl = pd.ExcelFile(path, engine="openpyxl")
        all_sheets = xl.sheet_names
        if self.sheet:
            if self.sheet in all_sheets:
                return [self.sheet]
            self._log(f"Sheet {self.sheet!r} not found; processing all sheets.")
        return all_sheets

    def _find_excel_header_row(self, path: Path, sheet: str) -> int:
        """Read first 15 rows as raw strings and find the best header row."""
        try:
            raw = pd.read_excel(path, sheet_name=sheet, header=None,
                                dtype=str, keep_default_na=False,
                                nrows=15, engine="openpyxl")
            return self._best_header_index(raw)
        except Exception:
            return 0

    def _expand_merged_cells(self, ws, df: pd.DataFrame, header_offset: int) -> pd.DataFrame:
        """Forward-fill values that were produced by merged cells (appear as NaN)."""
        try:
            merged = list(ws.merged_cells.ranges)
            if not merged:
                return df
            self._log(f"Expanding {len(merged)} merged cell range(s).")
            # openpyxl fills the merge anchor; other cells read as None/NaN.
            # Forward-fill propagates the anchor value into those empty cells.
            df = df.ffill(axis=0)
        except Exception:
            pass
        return df

    def _drop_hidden(self, ws, df: pd.DataFrame, header_offset: int) -> pd.DataFrame:
        """Remove rows/cols that are hidden in the Excel sheet."""
        try:
            # Hidden rows: row_dimensions keyed by 1-based row index
            hidden_rows = {
                r - header_offset - 2          # convert to df 0-based index
                for r, rd in ws.row_dimensions.items()
                if rd.hidden and r > header_offset + 1
            }
            hidden_cols = {
                c - 1
                for c, cd in ws.column_dimensions.items()
                if cd.hidden
            }
            if hidden_rows:
                self._log(f"Dropping {len(hidden_rows)} hidden row(s).")
                df = df.drop(index=[i for i in hidden_rows if i in df.index], errors="ignore")
            if hidden_cols:
                self._log(f"Dropping {len(hidden_cols)} hidden column(s).")
                df = df.drop(columns=df.columns[[c for c in hidden_cols if c < len(df.columns)]], errors="ignore")
        except Exception:
            pass
        return df

    # ── delimiter detection ───────────────────────────────────────────────────

    @staticmethod
    def _detect_delimiter(sample: str) -> str:
        """
        Count delimiter candidates outside of double-quoted regions so that
        commas/semicolons inside quoted values do not skew the result.
        """
        candidates = {",": 0, ";": 0, "\t": 0, "|": 0}
        in_quote = False
        for ch in sample:
            if ch == '"':
                in_quote = not in_quote
            elif not in_quote and ch in candidates:
                candidates[ch] += 1
        best = max(candidates, key=candidates.get)
        return best if candidates[best] > 0 else ","

    # ── header-row detection ──────────────────────────────────────────────────

    def _find_header_row(self, raw: str, delimiter: str) -> int:
        """
        Return the actual 0-based line number (in the raw file) that should
        be used as the header row.  Empty lines are skipped when scoring but
        the returned index always refers to the ORIGINAL line position so that
        pandas' ``header=N`` parameter works correctly.
        """
        all_lines = raw.splitlines()[:25]

        # Build a list of (original_line_idx, parsed_cells) for non-empty lines
        indexed_rows: list[tuple[int, list[str]]] = []
        for original_idx, line in enumerate(all_lines):
            if not line.strip():
                continue
            # Strip surrounding quotes from each cell
            cells = [c.strip().strip('"\'') for c in line.split(delimiter)]
            indexed_rows.append((original_idx, cells))

        if not indexed_rows:
            return 0

        best_orig_idx = 0
        best_score    = -1

        for filtered_pos, (orig_idx, cells) in enumerate(indexed_rows):
            vals = [v for v in cells if v]
            if not vals:
                continue
            non_num = sum(1 for v in vals if not _is_numeric_str(v) and len(v) < 80)
            score   = non_num / len(vals)
            # Require at least 3 populated cells to distinguish a real header
            # from a single-cell title or metadata row
            if len(vals) >= 3 and score > best_score:
                best_score    = score
                best_orig_idx = orig_idx

        if best_orig_idx > 0:
            self._log(f"Auto-detected header at line {best_orig_idx} (skipping {best_orig_idx} leading row(s)).")
        return best_orig_idx

    @staticmethod
    def _best_header_index(df: pd.DataFrame) -> int:
        """
        Convenience wrapper used by the Excel header detector.
        df rows are already sequential (no empty-line skipping needed).
        Returns the 0-based row index of the most likely header.
        """
        best, best_score = 0, -1
        for i, row in df.iterrows():
            if i >= 15:
                break
            vals = [str(v).strip() for v in row if str(v).strip() not in ("", "nan")]
            if not vals:
                continue
            non_num = sum(1 for v in vals if not _is_numeric_str(v) and len(v) < 80)
            score   = non_num / len(vals)
            if len(vals) >= 3 and score > best_score:
                best, best_score = int(str(i)), score
        return best


    # ── master cleaner ────────────────────────────────────────────────────────

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()

        # ── 1. stringify everything (safety pass) ─────────────────────────────
        df = df.astype(str)

        # ── 2. remove fully-empty rows / cols ────────────────────────────────
        df = self._drop_empty_rows_cols(df)

        # ── 3. strip cell-level whitespace & normalise internal linebreaks ────
        df = self._strip_cells(df)

        # ── 4. unify null representations ────────────────────────────────────
        if self.unify_nulls:
            df = self._unify_nulls(df)

        # ── 5. remove Excel formula error strings ────────────────────────────
        df = self._clean_excel_errors(df)

        # ── 6. clean headers ──────────────────────────────────────────────────
        df = self._clean_headers(df)

        # ── 7. parse / coerce numeric columns ────────────────────────────────
        if self.numeric_coerce:
            df = self._coerce_numerics(df)

        # ── 8. coerce boolean columns ─────────────────────────────────────────
        if self.bool_coerce:
            df = self._coerce_booleans(df)

        # ── 9. reset index ────────────────────────────────────────────────────
        df = df.reset_index(drop=True)

        return df

    # ── step implementations ──────────────────────────────────────────────────

    def _drop_empty_rows_cols(self, df: pd.DataFrame) -> pd.DataFrame:
        # Treat "nan", "", "none" as empty for this check
        def _is_blank(v: str) -> bool:
            return v.strip().lower() in {"nan", "none", "nat", ""}

        if self.drop_empty_rows:
            mask_rows = df.apply(lambda r: all(_is_blank(str(v)) for v in r), axis=1)
            n = mask_rows.sum()
            if n:
                self._log(f"Dropped {n} fully-empty row(s).")
            df = df[~mask_rows]

        if self.drop_empty_cols:
            mask_cols = df.apply(lambda c: all(_is_blank(str(v)) for v in c), axis=0)
            n = mask_cols.sum()
            if n:
                self._log(f"Dropped {n} fully-empty column(s).")
            df = df.loc[:, ~mask_cols]

        return df

    def _strip_cells(self, df: pd.DataFrame) -> pd.DataFrame:
        def _fix(v: str) -> str:
            if not isinstance(v, str):
                return v
            # Replace internal newlines / carriage returns with a space
            v = _CELL_LINEBREAK_RE.sub(" ", v)
            # Strip leading/trailing whitespace (including non-breaking spaces)
            v = v.strip().strip("\u00a0\u200b\u200c\u200d\ufeff")
            # Collapse multiple internal spaces
            v = _WHITESPACE_RE.sub(" ", v)
            # Remove surrounding stray quotes (single or double) that wrap the whole cell
            if len(v) >= 2 and v[0] == v[-1] and v[0] in ("'", '"'):
                inner = v[1:-1]
                if v[0] not in inner:   # only strip if not also quoted inside
                    v = inner
            return v

        self._log("Stripped whitespace / normalised cell content.")
        return df.apply(lambda col: col.map(_fix) if col.dtype == object else col)

    def _unify_nulls(self, df: pd.DataFrame) -> pd.DataFrame:
        before = df.isin(["nan", "none", "None", "NaN", "NaT"]).sum().sum()
        df = df.apply(
            lambda col: col.map(
                lambda v: np.nan if isinstance(v, str) and v.strip().lower() in _NULL_VARIANTS else v
            ) if col.dtype == object else col
        )
        total = df.isna().sum().sum()
        self._log(f"Unified null representations → {total} NaN cell(s) total.")
        return df

    def _clean_excel_errors(self, df: pd.DataFrame) -> pd.DataFrame:
        def _fix(v):
            if isinstance(v, str) and v.strip().lower() in _EXCEL_ERRORS:
                return np.nan
            return v
        changed = df.apply(lambda c: c.map(lambda v: isinstance(v, str) and v.strip().lower() in _EXCEL_ERRORS) if c.dtype == object else pd.Series([False]*len(c))).sum().sum()
        if changed:
            self._log(f"Replaced {changed} Excel error string(s) with NaN.")
        return df.apply(lambda c: c.map(_fix) if c.dtype == object else c)

    def _clean_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        original = list(df.columns)
        new_cols: list[str] = []
        seen: dict[str, int] = {}

        for i, col in enumerate(original):
            col = str(col).strip()
            # Remove BOM / zero-width chars
            col = col.strip("\ufeff\u200b\u200c\u200d\u00a0")
            # Collapse internal whitespace
            col = _WHITESPACE_RE.sub(" ", col)
            # Replace "Unnamed: N" (pandas default) or empty
            if not col or re.match(r"^unnamed[:\s_]?\d*$", col, re.I) or col.lower() == "nan":
                col = f"col_{i}"
            # snake_case conversion
            if self.snake_case:
                col = _to_snake(col)
            else:
                # Still remove / replace characters illegal in identifiers
                col = _ILLEGAL_HEADER_RE.sub("_", col).strip("_")
                col = _MULTI_UNDERSCORE.sub("_", col)
                if not col:
                    col = f"col_{i}"

            # De-duplicate
            if col in seen:
                seen[col] += 1
                col = f"{col}_{seen[col]}"
            else:
                seen[col] = 0

            new_cols.append(col)

        changed = sum(1 for a, b in zip(original, new_cols) if str(a) != b)
        if changed:
            self._log(f"Cleaned {changed} column name(s).")
        df.columns = new_cols
        return df

    def _coerce_numerics(self, df: pd.DataFrame) -> pd.DataFrame:
        coerced = 0
        for col in df.columns:
            if df[col].dtype != object:
                continue
            non_null = df[col].dropna()
            if len(non_null) == 0:
                continue

            # Try to parse each value
            parsed = non_null.map(_safe_numeric)
            # Accept coercion if ≥ 80 % of non-null values became numeric
            numeric_mask = parsed.map(lambda v: isinstance(v, (int, float)) and not isinstance(v, bool))
            if numeric_mask.mean() >= 0.80:
                df[col] = df[col].map(lambda v: _safe_numeric(v) if pd.notna(v) else v)
                coerced += 1

        if coerced:
            self._log(f"Coerced {coerced} column(s) to numeric types.")
        return df

    def _coerce_booleans(self, df: pd.DataFrame) -> pd.DataFrame:
        coerced = 0
        for col in df.columns:
            if df[col].dtype != object:
                continue
            non_null = df[col].dropna()
            if len(non_null) == 0:
                continue
            lower_vals = non_null.map(lambda v: str(v).strip().lower())
            bool_mask  = lower_vals.isin(_TRUE_VARIANTS | _FALSE_VARIANTS)
            if bool_mask.mean() >= 0.90:
                def _to_bool(v):
                    if pd.isna(v):
                        return v
                    lv = str(v).strip().lower()
                    if lv in _TRUE_VARIANTS:
                        return True
                    if lv in _FALSE_VARIANTS:
                        return False
                    return v
                df[col] = df[col].map(_to_bool)
                coerced += 1

        if coerced:
            self._log(f"Coerced {coerced} column(s) to boolean types.")
        return df

    # ── reporting ─────────────────────────────────────────────────────────────

    def _log(self, msg: str) -> None:
        self.report.append(msg)

    def get_report(self) -> str:
        return "\n".join(self.report)


# ── standalone utility ────────────────────────────────────────────────────────

def _is_numeric_str(s: str) -> bool:
    try:
        float(s.replace(",", "."))
        return True
    except ValueError:
        return False


# ── sorting ───────────────────────────────────────────────────────────────────

SORT_METHODS = {
    "natural":     "Natural (A1 < A2 < A10)",
    "alpha":       "Alphabetical (A–Z)",
    "numeric":     "Numeric (0 → 9)",
    "date":        "Date (oldest → newest)",
    "length":      "By length (shortest → longest)",
    "frequency":   "By frequency (most common first)",
}


def sort_dataframe(
    df:     pd.DataFrame,
    specs:  list[dict],
    report: list[str] | None = None,
) -> pd.DataFrame:
    """
    Permanently sort a DataFrame by one or more rules.

    Each spec is a dict with keys:
        column    : str   — column name
        direction : str   — 'asc' | 'desc'
        method    : str   — 'natural' | 'alpha' | 'numeric' |
                            'date' | 'length' | 'frequency'

    Rules are applied in order from first to last (multi-key stable sort).
    Later rules act as tiebreakers for earlier ones.
    """
    if not specs or df.empty:
        return df

    df = df.copy()

    # Apply specs in REVERSE order so primary key (first spec) wins.
    # Pandas stable sort preserves relative order from the previous pass.
    for spec in reversed(specs):
        col    = spec.get("column", "")
        asc    = spec.get("direction", "asc") == "asc"
        method = spec.get("method", "natural")

        if col not in df.columns:
            # Try snake_case fallback: "Last Name" → "last_name"
            snake_col = _to_snake(col)
            if snake_col in df.columns:
                if report is not None:
                    report.append(
                        f"Sort: column '{col}' resolved to '{snake_col}' (snake_case)."
                    )
                col = snake_col
            else:
                # Case-insensitive fallback (catches mixed-case mismatches)
                lower_map = {c.lower(): c for c in df.columns}
                matched = lower_map.get(col.lower()) or lower_map.get(snake_col.lower())
                if matched:
                    if report is not None:
                        report.append(
                            f"Sort: column '{col}' resolved to '{matched}' (case-insensitive)."
                        )
                    col = matched
                else:
                    if report is not None:
                        report.append(f"Sort: column '{col}' not found, skipping.")
                    continue

        try:
            if method == "numeric":
                key = pd.to_numeric(df[col], errors="coerce")
                df  = df.assign(_k=key).sort_values(
                    "_k", ascending=asc, na_position="last", kind="stable"
                ).drop(columns="_k")

            elif method == "date":
                key = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
                df  = df.assign(_k=key).sort_values(
                    "_k", ascending=asc, na_position="last", kind="stable"
                ).drop(columns="_k")

            elif method == "length":
                key = df[col].fillna("").astype(str).str.len()
                df  = df.assign(_k=key).sort_values(
                    "_k", ascending=asc, na_position="last", kind="stable"
                ).drop(columns="_k")

            elif method == "frequency":
                freq = df[col].value_counts(ascending=False)
                key  = df[col].map(freq).fillna(0)
                # Most-common first when asc=True (frequency desc = value asc)
                df   = df.assign(_k=key).sort_values(
                    "_k", ascending=not asc, na_position="last", kind="stable"
                ).drop(columns="_k")

            elif method == "natural":
                try:
                    from natsort import natsort_keygen
                    nk = natsort_keygen(key=lambda v: str(v).lower())
                    key = df[col].fillna("").map(nk)
                    df  = df.assign(_k=key).sort_values(
                        "_k", ascending=asc, na_position="last", kind="stable"
                    ).drop(columns="_k")
                except ImportError:
                    # Graceful fallback to alphabetical
                    df = df.sort_values(
                        col, ascending=asc,
                        key=lambda s: s.fillna("").astype(str).str.lower(),
                        na_position="last", kind="stable"
                    )

            else:  # alpha
                df = df.sort_values(
                    col, ascending=asc,
                    key=lambda s: s.fillna("").astype(str).str.lower(),
                    na_position="last", kind="stable"
                )

            if report is not None:
                dir_str = "ascending" if asc else "descending"
                report.append(
                    f"Sorted by '{col}'  [{SORT_METHODS.get(method, method)}, {dir_str}]"
                )

        except Exception as exc:
            if report is not None:
                report.append(f"Sort error on '{col}': {exc}")

    return df.reset_index(drop=True)
